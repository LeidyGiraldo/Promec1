from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from core.models import Categoria, Productos, Productos_produccion, Tipo_producto, Ubicacion
from calificacion.models import Calificacion
from citas.models import Cita

def Index(request):

    calificacio = Calificacion.objects.all()
    prodct = Productos.objects.all()
    return render(request, 'index.html', { 'producto': prodct, 'calificacion':calificacio })
    
@login_required    
def Ad_servicios(request):
    return render(request, 'ad_servicios.html', {
        
    })
    
def Contacto(request):
    return render(request, 'contacto.html', {
        
    })
    
@login_required
def Inventario(request):
    
    tipo = Tipo_producto.objects.all()
    ubicacion = Ubicacion.objects.all()
    producto = Productos.objects.all()
    return render(request, 'inventario.html', {'tipo': tipo, 'ubicacion': ubicacion, 'producto': producto})
    
def Login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            messages.success(request, 'Bienvenido')
            return redirect('index')
        else:
            messages.error(request, 'Credenciales incorrectas')
    return render(request, 'login.html',{
    })
    
def login_view(request):
    if request.user.is_authenticated:
        # Si el usuario ya está autenticado, cerrar la sesión
        logout(request)
    # Continuar con el manejo normal de la vista de inicio de sesión
    # ... 
    
def registro_view(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        username = request.POST.get('usernameR')
        password = request.POST.get('passwordR')
        user = User.objects.create_user(username=username, first_name=name, email=email, password=password)
        user.save()
        
        login(request, user)
        return redirect('index')

    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    messages.success(request, 'Sesión finalizada')
    return redirect('index')


from django.contrib.auth.decorators import login_required

@login_required
def protected_view(request):
    # Tu lógica para la vista protegida
    ...

    
@login_required
def Produccion(request):
    
    produccion = Productos_produccion.objects.all()
    categoria = Categoria.objects.all()
    return render(request, 'produccion.html',{'produccion': produccion, 'categoria':categoria})
    
def Solicitar_cita(request):
    return render(request, 'Solicitarcita.html',{
        
    })
    
@login_required
def Crear_inventario(request):
    if request.method == 'POST':
        id_producto = request.POST.get('codigo')
        nombre_producto = request.POST.get('nombre')
        imagen_producto = request.FILES.get('imagen')
        fecha_ingreso = request.POST.get('fecha')
        nit_proveedor = request.POST.get('nit')
        existencias = request.POST.get('existencias')
        descripcion = request.POST.get('descripcion')
        tipo_id = request.POST.get('tipo')
        ubicacion_id = request.POST.get('ubicacion')
        productos = Productos(id=id_producto, nombre_producto=nombre_producto, imagen_producto=imagen_producto, fecha_ingreso=fecha_ingreso, nit_proveedor=nit_proveedor, existencias=existencias, descripcion=descripcion, tipo_id=tipo_id, ubicacion_id=ubicacion_id )
        productos.save()
        return redirect('inventario')
    
    tipo = Tipo_producto.objects.all()
    ubicacion = Ubicacion.objects.all()
    return render(request, 'crear_producto_i.html',{'tipo': tipo, 'ubicacion': ubicacion})
    
@login_required
def Editar_producto(request, id):
    product = Productos.objects.get(pk= id)
    if request.method == 'POST':
        product.nombre_producto = request.POST.get('nombre')
        product.existencias = request.POST.get('existencias')
        
        product.save() 
        return redirect('inventario')
        
    return render(request, 'editar_producto.html', {'product':product})



@login_required
def Eliminar_producto(request, id):
    producto = get_object_or_404(Productos, id=id)
    
    if request.method == 'POST':
        producto.delete()
        return redirect('inventario')
    
    return render(request, 'eliminar_producto.html',{'producto': producto})

@login_required
def Crear_produccion(request):
    if request.method == 'POST':
        id_produccion = request.POST.get('codigo')
        nombre_producto = request.POST.get('nombre')
        cantidad = request.POST.get('cantidad')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_entrega = request.POST.get('fecha_entrega')
        duracion_estimada = request.POST.get('duracion')
        categoria = request.POST.get('categoria')
        
        Productos_p = Productos_produccion(id=id_produccion, nombre_producto=nombre_producto, cantidad=cantidad , fecha_inicio=fecha_inicio, fecha_entrega=fecha_entrega, duracion_estimada=duracion_estimada, categoria_id=categoria)
        Productos_p.save()
        return redirect('produccion')
    
    cat = Categoria.objects.all()
    return render(request, 'crear_produccion.html', {'categoria':cat})

@login_required
def Editar_produccion(request, id):
    product = Productos_produccion.objects.get(pk= id)
    if request.method == 'POST':
        product.cantidad = request.POST.get('cantidad')
        product.fecha_entrega = request.POST.get('fecha_entrega')
        product.duracion_estimada = request.POST.get('duracion')
        
        product.save()
        return redirect('produccion')
    return render(request, 'editar_produccion.html', {'produccion':product})

@login_required
def Eliminar_produccion(request, id):
    produccion= get_object_or_404(Productos_produccion, id=id)
    if request.method == 'POST':
        produccion.delete()
        return redirect('produccion')
    return render(request, 'eliminar_produccion.html', {'produccion': produccion})

@login_required
def Calificacion_view(request, id):
    user = User.objects.get(pk=id)
    if request.method == 'POST':
        contenido = request.POST.get('contenido')
        puntuacion = request.POST.get('puntuacion')
        usuario = user
        calificacion = Calificacion(contenido=contenido, puntuacion=puntuacion, usuario=usuario)
        calificacion.save()
        return redirect('index')
    return render(request, 'calificacion.html',{'user':user})

@login_required
def Citas(request):
    
    citas = Cita.objects.all()
    return render(request, 'citas.html',{'citas':citas})

@login_required
def eliminar_cita(request, id):
    cita = get_object_or_404(Cita, id=id)
    if request.method == 'POST':
        cita.delete()
        return redirect('citas')
    return render(request, 'eliminar_cita.html', {'citas':cita})


def Solicitar_cita(request):
    if request.method == 'POST':
        nombre_persona = request.POST.get('nombre')
        email = request.POST.get('correo')
        telefono = request.POST.get('telefono')
        fecha_cita = request.POST.get('fecha')
        hora_cita = request.POST.get('hora')
        servicio = request.POST.get('servicio')
        
        cita = Cita(nombre_persona=nombre_persona, email=email, telefono=telefono, fecha_cita=fecha_cita, hora_cita=hora_cita, servicio_id=servicio)
        cita.save()
        return redirect('index')
    
    return render(request, 'Solicitarcita.html',{})

def Generar_excel_inventario(request):
    inventario = Productos.objects.all()
    
    libro = Workbook()
    hoja = libro.active
    hoja.title = 'Reporte de inventario'
    
    border_style = Side(style='thin', color='000000')
    
    encabezados = ['nombre del proucto', 'tipo', 'fecha de ingreso', 'nit del proveedor', 'existencias', 'descripcion', 'ubicacion']
    hoja.append(encabezados)
    for fila in hoja.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(encabezados)):
        for celda in fila:
            celda.font = Font(bold=True, color='ffffff')
            celda.alignment = Alignment(horizontal='center')
            celda.fill = PatternFill(start_color="c31f1f", end_color="c31f1f", fill_type="solid")
            celda.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    for columna in range(2, len(encabezados) + 1): 
        for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=columna, max_col=columna):
            for celda in fila:
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal='center')
                celda.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                celda.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    for inventario in inventario:
        hoja.append([inventario.nombre_producto, inventario.tipo.nombre_tipo, inventario.fecha_ingreso, inventario.nit_proveedor, inventario.existencias, inventario.descripcion, inventario.ubicacion.nombre_ubicacion])
        
        for columna in hoja.columns:
            max_length = 0
            column = columna[0].column_letter
            for celda in columna:
                try:
                    if len(str(celda.value)) > max_length:
                        max_length = len(celda.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            hoja.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_inventario.xlsx'
    libro.save(response)
    
    return response

def General_excel_produccion(request):
    produccion = list(Productos_produccion.objects.all())
    
    libro = Workbook()
    hoja = libro.active
    hoja.title = 'Reporte de produccion'
    
    border_style = Side(style='thin', color='000000')
    
    encabezados = ['Nombre del producto', 'Cantidad', 'Fecha de inicio', 'Fecha de entrega', 'Duración estimada', 'Categoría']
    hoja.append(encabezados)
    for fila in hoja.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(encabezados)):
        for celda in fila:
            celda.font = Font(bold=True, color='ffffff')
            celda.alignment = Alignment(horizontal='center')
            celda.fill = PatternFill(start_color="c31f1f", end_color="c31f1f", fill_type="solid")
            celda.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    
    for columna in range(2, len(encabezados) + 1): 
        for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=columna, max_col=columna):
            for celda in fila:
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal='center')
                celda.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                celda.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    for p in produccion:
        hoja.append([p.nombre_producto, p.cantidad, p.fecha_inicio, p.fecha_entrega, p.duracion_estimada, p.categoria.nombre_categoria])
    
    for columna in hoja.columns:
        max_length = 0
        column = columna[0].column_letter
        for celda in columna:
            try:
                if len(str(celda.value)) > max_length:
                    max_length = len(celda.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        hoja.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_produccion.xlsx'
    libro.save(response)
    
    return response



def General_excel_citas(request):
    citas = list(Cita.objects.all())
    
    libro = Workbook()
    hoja = libro.active
    hoja.title = 'Reporte de citas'
    
    border_style = Side(style='thin', color='000000')
    
    encabezados = ['Nombre del cliente', 'Correo electrónico', 'telefono', 'Fecha de la cita', 'Hora de la cita', 'Servicio']
    hoja.append(encabezados)
    for fila in hoja.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(encabezados)):
        for celda in fila:
            celda.font = Font(bold=True, color='ffffff')
            celda.alignment = Alignment(horizontal='center')
            celda.fill = PatternFill(start_color="c31f1f", end_color="c31f1f", fill_type="solid")
            celda.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    
    for columna in range(2, len(encabezados) + 1): 
        for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=columna, max_col=columna):
            for celda in fila:
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal='center')
                celda.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                celda.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    for citas in citas:
        hoja.append([citas.nombre_persona, citas.email, citas.telefono, citas.fecha_cita, citas.hora_cita,citas.servicio.nombre_servicio])
    
    for columna in hoja.columns:
        max_length = 0
        column = columna[0].column_letter
        for celda in columna:
            try:
                if len(str(celda.value)) > max_length:
                    max_length = len(celda.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        hoja.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_citas.xlsx'
    libro.save(response)
    
    return response

# En tu archivo views.py
from django.shortcuts import render
from django.views import View

from django.shortcuts import render
from django.views import View

class ClientesView(View):
    def get(self, request):
        return render(request, 'clientes.html')  # Asegúrate de tener el archivo 'clientes.html'

def calendario_view(request):
    return render(request, 'calendario.html')
from django.shortcuts import render

def mantenimiento_view(request):
    # Datos estáticos para demostración
    mantenimientos = [
        {'nombre': 'Mantenimiento A', 'descripcion': 'Descripción del mantenimiento A', 'fecha': '2024-09-16'},
        {'nombre': 'Mantenimiento B', 'descripcion': 'Descripción del mantenimiento B', 'fecha': '2024-09-17'},
        # Agrega más datos si es necesario
    ]
    return render(request, 'mantenimiento.html', {'mantenimientos': mantenimientos})


# myapp/views.py
from django.shortcuts import render

def notifications_view(request):
    return render(request, 'notifications.html')
from django.shortcuts import render

def pedidos_view(request):
    return render(request, 'pedidos.html')
from django.shortcuts import render

def productos_list(request):
    productos = [
        {'id': 1, 'nombre': 'Producto 1', 'descripcion': 'Descripción del Producto 1', 'precio': '10.00', 'fecha_adicion': '2024-09-17'},
        {'id': 2, 'nombre': 'Producto 2', 'descripcion': 'Descripción del Producto 2', 'precio': '20.00', 'fecha_adicion': '2024-09-16'},
        {'id': 3, 'nombre': 'Producto 3', 'descripcion': 'Descripción del Producto 3', 'precio': '30.00', 'fecha_adicion': '2024-09-15'},
        # Agrega más productos según necesites
    ]
    return render(request, 'productos_list.html', {'productos': productos})

# views.py
from django.shortcuts import render

def categorias_view(request):
    categorias = [
        {'nombre': 'Electrónica'},
        {'nombre': 'Ropa'},
        {'nombre': 'Hogar'},
        {'nombre': 'Juguetes'},
        # Agrega más categorías según necesites
    ]
    return render(request, 'categorias.html', {'categorias': categorias})
